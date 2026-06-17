#!/usr/bin/env python3
"""Regenerate i18n/strings.json from the workbook's Localization sheet."""
import json, os
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "database", "mouaz-db-schema-and-localization.xlsx")
OUT = os.path.join(ROOT, "i18n", "strings.json")

ws = load_workbook(XLSX)["Localization"]
data = {}
for r in range(5, ws.max_row + 1):
    key = ws.cell(row=r, column=1).value
    if not key:
        continue
    data[key] = {"en": ws.cell(row=r, column=2).value, "ar": ws.cell(row=r, column=3).value}

payload = {"_meta": {"languages": ["en", "ar"], "default": "en", "rtl": ["ar"], "count": len(data)},
           "strings": data}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(payload, f, ensure_ascii=False, indent=2)
print(f"wrote {OUT}: {len(data)} keys")
