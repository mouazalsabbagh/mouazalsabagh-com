from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY = "1F3A4D"; TEAL = "2E6E73"; ACCENT = "9ACD32"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor=TEAL)
ZEBRA = PatternFill("solid", fgColor="F2F5F7")
YELLOW = PatternFill("solid", fgColor="FFF6CC")
PROPOSED = PatternFill("solid", fgColor="FDEAD7")
WHITE_BOLD = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE = Font(name=FONT, bold=True, color=NAVY, size=16)
SUBTITLE = Font(name=FONT, italic=True, color="555555", size=10)
BOLD = Font(name=FONT, bold=True, size=11)
REG = Font(name=FONT, size=10)
MONO = Font(name="Courier New", size=9)
thin = Side(style="thin", color="C8D0D6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
L = Alignment(horizontal="left", vertical="center", wrap_text=True)
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
RTL = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)

wb = Workbook()

def style_header(ws, row, ncols, fill=HDR_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = WHITE_BOLD; cell.fill = fill
        cell.alignment = C; cell.border = BORDER

def table(ws, start_row, headers, rows, widths, rtl_cols=None, proposed=False):
    rtl_cols = rtl_cols or []
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))
    for i, rdata in enumerate(rows, 1):
        r = start_row + i
        for j, val in enumerate(rdata, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = REG; cell.border = BORDER
            cell.alignment = RTL if j in rtl_cols else L
            if i % 2 == 0:
                cell.fill = ZEBRA
            if proposed:
                cell.fill = PROPOSED
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = f"{get_column_letter(1)}{start_row}:{get_column_letter(len(headers))}{start_row+len(rows)}"
    return start_row + len(rows)

# ───────────────────────── 1. OVERVIEW ─────────────────────────
ws = wb.active; ws.title = "Overview"
ws.sheet_view.showGridLines = False
ws["A1"] = "Mouaz AlSabbagh — Database Schema & Localization"; ws["A1"].font = TITLE
ws["A2"] = "Content organized into a relational schema + English/Arabic localization layer.  Generated 2026-06-03."
ws["A2"].font = SUBTITLE
ws.merge_cells("A1:F1"); ws.merge_cells("A2:F2")

ws["A4"] = "Workbook contents"; ws["A4"].font = BOLD
idx = [
    ("recommendations", "LIVE table — recommendation-letter submissions (from rec/collect.php)"),
    ("pages", "PROPOSED table — template-driven page creation / draft workflow"),
    ("translations", "PROPOSED table — UI string i18n (key → en/ar)"),
    ("Lookups", "Allowed values (ENUMs) and suggested option lists"),
    ("Localization", "English ⇄ Arabic UI strings, ready to drive multi-language display"),
    ("DDL", "Copy-paste CREATE TABLE statements"),
]
table(ws, 5, ["Sheet", "Description"], idx, [22, 92])

# Live summary (formulas — recalculated, never hardcoded)
sr = 5 + len(idx) + 2
ws.cell(row=sr, column=1, value="Live metrics").font = BOLD
# Data rows start at row 5 on every schema/localization sheet (header at row 4)
metrics = [
    ("Columns in 'recommendations'", "=COUNTA(recommendations!A5:A1000)"),
    ("Columns in 'pages' (proposed)", "=COUNTA(pages!A5:A1000)"),
    ("Columns in 'translations' (proposed)", "=COUNTA(translations!A5:A1000)"),
    ("Localization strings defined", "=COUNTA(Localization!A5:A1000)"),
    ("Strings missing Arabic", "=COUNTA(Localization!A5:A1000)-COUNTA(Localization!C5:C1000)"),
    ("Localization coverage (Arabic)", "=IF(COUNTA(Localization!A5:A1000)=0,0,COUNTA(Localization!C5:C1000)/COUNTA(Localization!A5:A1000))"),
]
table(ws, sr + 1, ["Metric", "Value"], [[m[0], None] for m in metrics], [40, 18])
for i, (_, f) in enumerate(metrics, 1):
    cell = ws.cell(row=sr + 1 + i, column=2, value=f)
    cell.font = Font(name=FONT, size=10, color="000000"); cell.alignment = C
ws.cell(row=sr + 1 + 6, column=2).number_format = "0.0%"  # coverage as %

leg = sr + 1 + len(metrics) + 2
ws.cell(row=leg, column=1, value="Legend").font = BOLD
ws.cell(row=leg+1, column=1, value="Orange rows = PROPOSED (not yet created in MySQL).").font = REG
ws.cell(row=leg+1, column=1).fill = PROPOSED
ws.cell(row=leg+2, column=1, value="Yellow = needs your input / update.").font = REG
ws.cell(row=leg+2, column=1).fill = YELLOW

# ───────────────────────── 2. recommendations ─────────────────────────
ws = wb.create_sheet("recommendations")
ws.sheet_view.showGridLines = False
ws["A1"] = "Table: recommendations  (LIVE)"; ws["A1"].font = TITLE
ws["A2"] = "Engine InnoDB · CHARSET utf8mb4 · auto-created by installTable() in rec/collect.php"; ws["A2"].font = SUBTITLE
hdr = ["Column", "Data Type", "Null", "Default", "Key", "Description"]
rows = [
 ["id","INT","NO","AUTO_INCREMENT","PK","Primary key"],
 ["submitted_at","DATETIME","NO","CURRENT_TIMESTAMP","","Time submission received"],
 ["ip_address","VARCHAR(45)","YES","NULL","","Submitter IP (IPv4/IPv6)"],
 ["rec_name","VARCHAR(200)","YES","NULL","","Recommender full name"],
 ["rec_title","VARCHAR(200)","YES","NULL","","Recommender job title"],
 ["rec_company","VARCHAR(200)","YES","NULL","","Recommender company"],
 ["rec_email","VARCHAR(200)","YES","NULL","","Recommender email"],
 ["rec_contact","VARCHAR(200)","YES","NULL","","Phone or LinkedIn URL"],
 ["rec_date","VARCHAR(30)","YES","NULL","","Letter date as entered"],
 ["rel_type","VARCHAR(60)","YES","NULL","","Relationship type (see Lookups)"],
 ["rel_duration","VARCHAR(30)","YES","NULL","","Relationship duration"],
 ["rel_context","TEXT","YES","NULL","","How they worked together"],
 ["target_role","VARCHAR(100)","YES","NULL","","Role being applied for"],
 ["target_company","VARCHAR(200)","YES","NULL","","Target company"],
 ["target_industry","VARCHAR(100)","YES","NULL","","Target industry"],
 ["strengths","TEXT","YES","NULL","","Key strengths"],
 ["obs_project","TEXT","YES","NULL","","Project observation"],
 ["obs_character","TEXT","YES","NULL","","Character observation"],
 ["letter_tone","VARCHAR(60)","YES","NULL","","Desired tone (see Lookups)"],
 ["letter_length","VARCHAR(30)","YES","NULL","","Desired length (see Lookups)"],
 ["lang","VARCHAR(5)","YES","'en'","IDX","Letter language: en | ar"],
 ["generated_letter","LONGTEXT","YES","NULL","","AI-generated letter text"],
 ["status","ENUM('new','reviewed','downloaded')","YES","'new'","","Workflow status"],
]
end = table(ws, 4, hdr, rows, [18, 34, 7, 20, 7, 40])
# highlight the multilingual driver column
for r in range(5, end + 1):
    if ws.cell(row=r, column=1).value == "lang":
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = YELLOW
        ws.cell(row=r, column=6).comment = Comment(
            "Drives which language the letter is generated in. Mirrors the 'translations' / Localization layer (en/ar).", "schema")

# ───────────────────────── 3. pages (proposed) ─────────────────────────
ws = wb.create_sheet("pages")
ws.sheet_view.showGridLines = False
ws["A1"] = "Table: pages  (PROPOSED)"; ws["A1"].font = TITLE
ws["A2"] = "Backs the template-driven page creation / edit / draft workflow (NEXT-STEPS section B)."; ws["A2"].font = SUBTITLE
rows = [
 ["id","INT","NO","AUTO_INCREMENT","PK","Primary key"],
 ["slug","VARCHAR(160)","NO","","UNIQUE","URL slug, kebab-case"],
 ["type","ENUM('case-study','project-preview','page')","NO","'page'","","Page kind"],
 ["status","ENUM('draft','published','archived')","NO","'draft'","IDX","Publish state"],
 ["lang","ENUM('en','ar')","NO","'en'","IDX","Content language"],
 ["title","VARCHAR(255)","YES","NULL","","Page / case-study title"],
 ["og_image","VARCHAR(255)","YES","NULL","","Social share image path (.webp)"],
 ["body_html","LONGTEXT","YES","NULL","","Rendered page body"],
 ["meta_json","JSON","YES","NULL","","Flexible fields: metrics, tags, client"],
 ["created_at","DATETIME","NO","CURRENT_TIMESTAMP","","Created"],
 ["updated_at","DATETIME","NO","CURRENT_TIMESTAMP ON UPDATE","","Last edited"],
]
table(ws, 4, hdr, rows, [16, 40, 7, 26, 8, 38], proposed=True)
ws.cell(row=4, column=5).comment = Comment("Add a composite UNIQUE index on (slug, lang) so EN and AR versions of the same page coexist.", "schema")

# ───────────────────────── 4. translations (proposed) ─────────────────────────
ws = wb.create_sheet("translations")
ws.sheet_view.showGridLines = False
ws["A1"] = "Table: translations  (PROPOSED)"; ws["A1"].font = TITLE
ws["A2"] = "UI string internationalization. One row per key; columns per language."; ws["A2"].font = SUBTITLE
rows = [
 ["id","INT","NO","AUTO_INCREMENT","PK","Primary key"],
 ["string_key","VARCHAR(120)","NO","","UNIQUE","Dot key, e.g. nav.home"],
 ["en","TEXT","YES","NULL","","English string"],
 ["ar","TEXT","YES","NULL","","Arabic string"],
 ["context","VARCHAR(160)","YES","NULL","","Where it is used"],
 ["updated_at","DATETIME","NO","CURRENT_TIMESTAMP ON UPDATE","","Last edited"],
]
table(ws, 4, hdr, rows, [16, 34, 7, 26, 8, 40], proposed=True)

# ───────────────────────── 5. Lookups ─────────────────────────
ws = wb.create_sheet("Lookups")
ws.sheet_view.showGridLines = False
ws["A1"] = "Lookups — allowed values & suggested options"; ws["A1"].font = TITLE
look = [
 ["recommendations.status","new | reviewed | downloaded","ENUM (enforced)"],
 ["pages.status","draft | published | archived","ENUM (proposed)"],
 ["pages.type","case-study | project-preview | page","ENUM (proposed)"],
 ["lang","en | ar","2-letter ISO code"],
 ["rel_type","Manager | Direct Report | Colleague | Client | Mentor | Professor","suggested"],
 ["letter_tone","Professional | Warm | Formal | Enthusiastic","suggested"],
 ["letter_length","Short | Medium | Long","suggested"],
]
table(ws, 3, ["Field", "Allowed / suggested values", "Note"], look, [28, 56, 22])

# ───────────────────────── 6. Localization (en/ar) ─────────────────────────
ws = wb.create_sheet("Localization")
ws.sheet_view.showGridLines = False
ws["A1"] = "Localization — English ⇄ Arabic UI strings"; ws["A1"].font = TITLE
ws["A2"] = "Seed data for the 'translations' table / a JSON i18n file. Arabic column is RTL."; ws["A2"].font = SUBTITLE
loc = [
 ["nav.home","Home","الرئيسية","Main navigation"],
 ["nav.about","About","نبذة عني","Main navigation"],
 ["nav.works","Works","الأعمال","Main navigation"],
 ["nav.projects","Projects","المشاريع","Main navigation"],
 ["nav.services","Services","الخدمات","Main navigation"],
 ["nav.blog","Blog","المدونة","Main navigation"],
 ["nav.contact","Contact","تواصل معي","Main navigation"],
 ["btn.view_cv","View CV","عرض السيرة الذاتية","Contact page"],
 ["btn.contact_me","Contact Me","تواصل معي","Hero / CTA"],
 ["hero.role","Senior Art Director & Creative Technologist","مدير فني أول وتقني إبداعي","Hero"],
 ["section.work_education","Work & Education","الخبرات والتعليم","About"],
 ["rec.title","Recommendation Letter Generator","منشئ خطابات التوصية","Recommendation form"],
 ["rec.generate","Generate Recommendation Letter","إنشاء خطاب التوصية","Recommendation form"],
 ["rec.copy","Copy to Clipboard","نسخ إلى الحافظة","Recommendation form"],
 ["rec.print","Print / Save PDF","طباعة / حفظ PDF","Recommendation form"],
 ["rec.edit","Edit & Regenerate","تعديل وإعادة الإنشاء","Recommendation form"],
 ["form.name","Your Name","الاسم","Forms"],
 ["form.email","Email","البريد الإلكتروني","Forms"],
 ["form.message","Message","الرسالة","Forms"],
 ["form.submit","Send Message","إرسال الرسالة","Forms"],
 ["status.new","New","جديد","Admin dashboard"],
 ["status.reviewed","Reviewed","تمت المراجعة","Admin dashboard"],
 ["status.downloaded","Downloaded","تم التنزيل","Admin dashboard"],
 ["footer.rights","All Rights Reserved","جميع الحقوق محفوظة","Footer"],
]
table(ws, 4, ["string_key", "en (English)", "ar (العربية)", "Context"], loc, [24, 44, 38, 22], rtl_cols=[3])

# ───────────────────────── 7. DDL ─────────────────────────
ws = wb.create_sheet("DDL")
ws.sheet_view.showGridLines = False
ws["A1"] = "DDL — CREATE TABLE statements"; ws["A1"].font = TITLE
ws.column_dimensions["A"].width = 110
ddl = """-- LIVE (already auto-created by rec/collect.php)
CREATE TABLE IF NOT EXISTS recommendations (
  id INT AUTO_INCREMENT PRIMARY KEY,
  submitted_at DATETIME DEFAULT CURRENT_TIMESTAMP,
  ip_address VARCHAR(45),
  rec_name VARCHAR(200), rec_title VARCHAR(200), rec_company VARCHAR(200),
  rec_email VARCHAR(200), rec_contact VARCHAR(200), rec_date VARCHAR(30),
  rel_type VARCHAR(60), rel_duration VARCHAR(30), rel_context TEXT,
  target_role VARCHAR(100), target_company VARCHAR(200), target_industry VARCHAR(100),
  strengths TEXT, obs_project TEXT, obs_character TEXT,
  letter_tone VARCHAR(60), letter_length VARCHAR(30),
  lang VARCHAR(5) DEFAULT 'en',
  generated_letter LONGTEXT,
  status ENUM('new','reviewed','downloaded') DEFAULT 'new',
  INDEX (lang), INDEX (status)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;

-- PROPOSED: template-driven pages
CREATE TABLE IF NOT EXISTS pages (
  id INT AUTO_INCREMENT PRIMARY KEY,
  slug VARCHAR(160) NOT NULL,
  type ENUM('case-study','project-preview','page') NOT NULL DEFAULT 'page',
  status ENUM('draft','published','archived') NOT NULL DEFAULT 'draft',
  lang ENUM('en','ar') NOT NULL DEFAULT 'en',
  title VARCHAR(255), og_image VARCHAR(255),
  body_html LONGTEXT, meta_json JSON,
  created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
  updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  UNIQUE KEY uq_slug_lang (slug, lang), INDEX (status), INDEX (lang)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;

-- PROPOSED: UI i18n strings
CREATE TABLE IF NOT EXISTS translations (
  id INT AUTO_INCREMENT PRIMARY KEY,
  string_key VARCHAR(120) NOT NULL UNIQUE,
  en TEXT, ar TEXT,
  context VARCHAR(160),
  updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;"""
for i, line in enumerate(ddl.split("\n"), 3):
    cell = ws.cell(row=i, column=1, value=line); cell.font = MONO; cell.alignment = Alignment(horizontal="left")

wb.save("database/mouaz-db-schema-and-localization.xlsx")
print("saved")
